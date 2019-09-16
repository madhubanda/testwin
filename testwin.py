import contextlib
import selenium.webdriver as webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
from openpyxl import Workbook
import datetime
import time
import itertools
import random
import re
from string import ascii_lowercase
import wx
import getpass
import os
from datetime import date

cutdate = '2019-09-12'

url = "http://www.gumtree.com.au"
email = "madhanreddyaus@gmail.com"
#email = "madanmohanreddy4b3@gmail.com"
password = "tvsapache1"

class Example(wx.Frame):
	
	def __init__(self, parent, title):
		super(Example, self).__init__(parent, title=title)
		self.InitUI()
		self.Centre()

	def InitUI(self):
		panel = wx.Panel(self)
		#self.isBrow = False
		font = wx.SystemSettings.GetFont(wx.SYS_SYSTEM_FONT)
		font.SetPointSize(9)
		vbox = wx.BoxSizer(wx.VERTICAL)
		hbox1 = wx.BoxSizer(wx.HORIZONTAL)
		vbox.Add((-1, 25))
		st1 = wx.StaticText(panel, label='Username')
		st1.SetFont(font)
		hbox1.Add(st1, flag=wx.RIGHT, border=10)
		self.tc1 = wx.TextCtrl(panel)
		hbox1.Add(self.tc1, 0, wx.ALL | wx.EXPAND, 5)
		vbox.Add(hbox1, border=10)
		vbox.Add((-1, 15))
		hbox2 = wx.BoxSizer(wx.HORIZONTAL)
		st1 = wx.StaticText(panel, label='Password')
		st1.SetFont(font)
		hbox2.Add(st1, flag=wx.RIGHT, border=10)
		self.tc2 = wx.TextCtrl(panel, style=wx.TE_PASSWORD)
		hbox2.Add(self.tc2, 0, wx.ALL | wx.EXPAND, 5)
		vbox.Add(hbox2, border=10)
		#change
		vbox.Add((0, 25))
		cb1 = wx.CheckBox(panel, label='Google Chrome')
		cb1.SetValue(False)
		cb1.Bind(wx.EVT_CHECKBOX, self.selectBrowser)
		vbox.Add(cb1,flag=wx.LEFT,border=30)
		vbox.Add((0, 25))
		cb2 = wx.CheckBox(panel, label='Firefox')
		cb2.SetValue(False)
		cb2.Bind(wx.EVT_CHECKBOX, self.selectBrowser)
		vbox.Add(cb2,flag=wx.LEFT,border=30)
		#change
		vbox.Add((0, 25))
		hbox3 = wx.BoxSizer(wx.HORIZONTAL)
		btn1 = wx.Button(panel, label='IMPORT 2 XL', size=(70, 30))
		hbox3.Add(btn1)
		#self.Bind(wx.EVT_BUTTON, self.Import2XL)
		btn2 = wx.Button(panel, label='AD UPDATE', size=(70, 30))
		hbox3.Add(btn2, flag=wx.LEFT, border=25)
		self.Bind(wx.EVT_BUTTON, self.Import2XL)
		vbox.Add(hbox3, flag=wx.ALIGN_LEFT|wx.RIGHT, border=10)
		panel.SetSizer(vbox)
	
	def selectBrowser(self,e):
		sender = e.GetEventObject()
		self.isBrow = sender.GetValue()
		self.isBrowLabel = sender.GetLabel()
		#print "Get checkbox label %s" % self.isBrowLabel
		return self.isBrowLabel
		
	def Import2XL(self,e):
		sender = e.GetEventObject()
		print('\n GUMTREE APP STARTS...')
		print ('Get button label'+" "+self.isBrowLabel)
		time.sleep(10)
		if sender.GetLabel() == 'IMPORT 2 XL':
			self.display("import")
		elif sender.GetLabel() == 'AD UPDATE':
			self.display("update")
		e.Skip()
	
	def display(self,action):
		print ("hello")
		if self.isBrowLabel == 'Google Chrome' or self.isBrowLabel == 'Firefox':
			print ("Browser is selected"+" "+self.isBrowLabel)
		else:
			print ("\n Found Browser is not selected, Please Try again")
			time.sleep(3000)
		uname = self.tc1.GetValue()
		pwd = self.tc2.GetValue()
		global cutdate
		print ("cutdate"+" "+cutdate)
		today = str(date.today())
		today = today.split("-")
		cutdate = str(cutdate).split("-")
		#print cutdate, today
		#print "event %s" % action
		#time.sleep(300)
		if cutdate[0] == today[0] and cutdate[1] == today[1] and cutdate[2] >= today[2]:
			driver = self.openBrowser(self.isBrowLabel)
			driver = self.loginApp(driver,uname,pwd)
			ws,wb = self.openXL("ads_list.xlsx")
			if action == "import":
				driver = self.readContent(driver,url,ws)
				wb.save("ads_list.xlsx")
			elif action == "update":
				driver = self.updateAds(driver,url,ws)
		else:
			print ("License expired")
			quit()

			
	def openBrowser(self,isBrowLabel):
		if self.isBrowLabel == 'Firefox':
			binary = FirefoxBinary('C:\\Program Files\\Mozilla Firefox\\firefox.exe')
			self.driver = webdriver.Firefox(firefox_binary=binary,executable_path=r'geckodriver.exe')
		elif self.isBrowLabel == 'Google Chrome':
			#driver = webdriver.Chrome(executable_path="C:\Python27\chromedriver.exe")
			try:
				self.driver = webdriver.Chrome(executable_path="chromedriver.exe")
			except:
				dirpath = os.getcwd()
				self.driver = webdriver.Chrome(executable_path=dirpath+"/"+"chromedriver")
		time.sleep(float(str(random.randrange(2,5,1))+"."+str(random.randrange(500,1000,12))))
		time.sleep(5)
		self.driver.maximize_window()
		return self.driver

	def loginApp(self,driver,uname,pwd):
		self.driver.get(url)
		time.sleep(float(str(random.randrange(2,5,1))+"."+str(random.randrange(500,1000,66))))
		self.driver.find_element_by_xpath('//*[@id=\"header-new\"]/div/div[1]/div[1]/div[1]/ul/li[5]').click()
		time.sleep(float(str(random.randrange(2,5,1))+"."+str(random.randrange(500,1000,36))))
		self.driver.find_element_by_xpath('//*[@id=\"login-email\"]').send_keys(uname)
		self.driver.find_element_by_xpath('//*[@id=\"login-password\"]').send_keys(pwd)
		time.sleep(float(str(random.randrange(1,5,1))+"."+str(random.randrange(100,1000,36))))
		self.driver.find_element_by_xpath('//*[@id=\"btn-submit-login\"]').send_keys(Keys.RETURN)
		time.sleep(float(str(random.randrange(1,4,1))+"."+str(random.randrange(200,1000,36))))
		return self.driver
	
	def openAds(self,driver):
		self.driver.find_element_by_xpath('//*[@id=\"nav-my\"]/div').click()
		self.driver.find_element_by_xpath('//*[@id=\"my-gumtree-nav\"]/li[3]/a').click()
		time.sleep(float(str(random.randrange(1,5,1))+"."+str(random.randrange(100,1000,36))))
		self.driver.find_element_by_xpath('//*[@id=\"my-nav\"]/div/ul/li[1]/ul/li[2]').click()
		self.titles = [];self.category =[];self.posted = [];self.links = [];self.repost = []
		scroll = 1500
		last_height = driver.execute_script("return document.body.scrollHeight")
		while True:
			self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
			time.sleep(3)
			new_height = self.driver.execute_script("return document.body.scrollHeight")
			if new_height == last_height:
				break
			else:
				last_height = new_height
		time.sleep(float(str(random.randrange(2,5,1))+"."+str(random.randrange(100,1000,36))))
		htm = self.driver.find_element_by_id("my-adlisting").get_attribute("innerHTML")
		self.titles = re.compile("data-ad-id=\"\S+\"\sdata-ad-title=\"([A-Za-z0-9\s+\-\,\:\&\;\.]+)+\"\sdata-render").findall(htm)
		self.category = re.compile("class=\"rs-ma-details-category\"\>([A-Za-z0-9\s+\-\,\:\&\;\.]+)+\<\/dd\>").findall(htm)
		self.posted = re.compile("\s+(\d+\/\d+\/\d+)\<\/dd\>").findall(htm)
		self.links = re.compile("\"rs-ad-title\"\s+href=\"\/([A-Za-z0-9\s+\-\:\&\?\=\.]+)+\"\>").findall(htm)
		self.repost = re.compile("\"repost-ad-free\s+j-track-link\"\s+href=\"\/(([A-Za-z0-9\s+\-\,\:\&\;\?\=\.]+)+)\"\s+data-tracker=").findall(htm)
		#print "\n Titles %s" % self.titles
		#print "\n Category %s" % self.category
		#print "\n Posted Date %s" % self.posted
		#print "\n Links %s" % self.links
		return self.driver,self.titles,self.category,self.posted,self.links,self.repost
	
	def readContent(self,driver,url,ws):
		self.driver,self.titles,self.category,self.posted,self.links,self.repost = self.openAds(self.driver)
		#print "Data collected"
		count = 0
		indx = 0
		for web in self.links:
			main_window = driver.current_window_handle
			self.driver.execute_script("window.open();")
			self.driver.switch_to_window(driver.window_handles[1])
			try:
				self.driver.get(url+"/"+web)
			except:
				sleep(2)
				self.driver.get(url+"/"+web)
			time.sleep(float(str(random.randrange(1,4,1))+"."+str(random.randrange(148,1000,36))))
			self.driver.execute_script("window.scrollTo(0, 900);")
			ad_perf = driver.find_element_by_id("ad-performance").get_attribute("innerHTML")		
			#print ad_perf
			views = re.compile("\"ad-performance-stats__value\"\>(\d+)\<\/span\>").findall(ad_perf)
			#print "views %s" % views[0]
			replies = re.compile("\"ad-performance-stats__replies\"\>\n\s+\<span\s+class=\"ad-performance-stats__value\"\>(\d+)\<\/span\>").findall(ad_perf)
			#print "replies %s" % views[1]
			try:
				htmcontent = self.driver.find_element_by_id('ad_description_details').get_attribute('innerHTML')
			except:
				try:
					htmcontent = self.driver.find_element_by_id('//*[@id=\"ad_description_details_content\"]').get_attribute('innerHTML')
				except:
					htmcontent = self.driver.find_element_by_css_selector('#ad_description_details_content').get_attribute('innerHTML')
			htmcontent = htmcontent.split("</div>")
			content = ""
			for line in htmcontent:
				line = line.strip()
				if line != "":
					if line[0] != "<":
						content = content+line
					#	print "\n line %s" % line
			#print "Content ==> %s" % content
			self.ws = self.writeXL(count,self.ws,self.titles[count],self.category[count],self.posted[indx],content,views[0],views[1])
			count = count+1
			indx = count+1
			self.driver.close()
			self.driver.switch_to_window(main_window)
		return self.driver,self.ws
		
	def updateAds(self,driver,url,ws):
		self.driver,self.titles,self.category,self.posted,self.links,self.repost = self.openAds(self.driver)
		for web in self.repost:
			main_window = driver.current_window_handle
			self.driver.execute_script("window.open();")
			self.driver.switch_to_window(driver.window_handles[1])
			#print "url ==>%s" % url
			#print "web ==>%s" % web
			try:
				self.driver.get(url+"/"+web)
			except:
				time.sleep(2)
				self.driver.get(url+"/"+web)
			time.sleep(float(str(random.randrange(1,4,1))+"."+str(random.randrange(148,1000,36))))			
			self.driver.find_element_by_xpath("//*[@id=\"modal-close-button-0\"]/svg").click()
			self.driver.close()
			self.driver.switch_to_window(main_window)
		return self.driver
	
	def openXL(self,filename):
		try:
			self.wb = load_workbook(filename)
		except:
			self.wb = Workbook()
			self.wb.create_sheet(filename)
		self.ws = self.wb.active
		self.ws['A1'] = "GUMTREE"
		self.ws['B1'] = "AD TITLE"
		self.ws['C1'] = "AD CATEGORY"
		self.ws['D1'] = "POSTED DATE"
		self.ws['E1'] = "AD VIEWS"
		self.ws['F1'] = "AD REPLIES"		
		self.ws['G1'] = "AD DESCRIPTION"
		return self.ws, self.wb
	
	def writeXL(self,indx,ws,titles,category,posted,content,views,replies):
		#print "%s,%s,%s"%(titles,category,posted)
		#print "spreadsheet active"
		#print "index is %s" % indx
		#print "A+indx %s" % 'A'+str(1+indx)
		self.ws['A'+str(2+indx)] = "GUMTREE"
		self.ws['B'+str(2+indx)] = titles
		#print "%s column populated" % str(1+indx)
		self.ws['C'+str(2+indx)] = category
		self.ws['D'+str(2+indx)] = posted
		self.ws['E'+str(2+indx)] = views
		self.ws['F'+str(2+indx)] = replies		
		self.ws['G'+str(2+indx)] = content
		#print "Total ads captured"
		return self.ws

def main():
	app = wx.App()
	ex = Example(None, title='GUMTREE UPDATE')
	ex.Show()
	app.MainLoop()
	
	
if __name__ == '__main__':
	main()
	
